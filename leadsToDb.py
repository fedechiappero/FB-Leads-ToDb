from openpyxl import load_workbook
import sys
from termcolor import colored
import argparse

parser = argparse.ArgumentParser(description='Generate insert queries from a xlsx file')
parser.add_argument('-f', '--file', help='xlsx file path')
parser.add_argument('-c', '--comment', help='Comment for records to insert')
args = parser.parse_args()

try:
    file = args.file
    if(file == None):
        raise Exception('Provide a path for xlsx file.')

    comment = args.comment
    if(comment == None):
        raise Exception('Provide a comment for records to insert.')
except Exception as e:
    print(colored(e, 'red'), f'Run {sys.argv[0]} -h for help')
    exit()

data = load_workbook(file)
dataSetLen = data['Datos'].max_row
insertContact = 'insert into contact (`name`, lastname, email, cellphone, id_cms_users, id_status, comments, created_at, updated_at) values ('
insertActivity = 'insert into activity (created_at, updated_at, id_contact, id_activity_type) values ('

rowsContact = []
rowsActivity = []

for row in data['Datos'].iter_rows(min_row=2):
    lastName = row[12].value.split()[len(row[12].value.split())-1]
    name = row[12].value if lastName == row[12].value else row[12].value.replace(lastName, '')
    email = row[14].value
    phone = str(row[13].value).strip()
    if phone[0] == '=':
        phone = phone.replace('=', '')
    cmsUser = '27' if row[12].row < dataSetLen/2 else '44'
    rowsContact.append((insertContact,f"'{name}',", f"'{lastName}',", f"'{email}',", f"'{phone}',", f"'{cmsUser}',", "'1',", f"'{comment}',", "now(),", "now())",";"))
        
for row in data['Datos'].iter_rows(min_row=2):
    rowsActivity.append((insertActivity,"now(),", "now(),","", ",'1')",";"))

sqlContact = data.create_sheet("Sql - Contact")
for row in rowsContact:
    sqlContact.append(row)

sqlActivity = data.create_sheet("Sql - Activity")
for row in rowsActivity:
    sqlActivity.append(row)

data.save(file)

print(colored('Success!', 'green'))